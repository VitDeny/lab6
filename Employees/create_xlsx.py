import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def calculate_age(birth_date_str):
    try:
        birth_date = datetime.strptime(birth_date_str, '%d.%m.%Y')
        today = datetime.now()
        age = today.year - birth_date.year
        if (today.month, today.day) < (birth_date.month, birth_date.day):
            age -= 1
        return age
    except:
        return None

def categorize_age(age):
    if age is None:
        return None
    if age < 18:
        return 'younger_18'
    elif age <= 45:
        return '18-45'
    elif age <= 70:
        return '45-70'
    else:
        return 'older_70'

def style_header(ws):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_xlsx_from_csv(csv_filename='employees.csv', xlsx_filename='employees.xlsx'):
    
    # Перевірка наявності CSV файлу
    if not os.path.exists(csv_filename):
        print(f"✗ Помилка: файл {csv_filename} не знайдено!")
        return
    
    try:
        # Зчитування даних з CSV
        employees_data = []
        with open(csv_filename, 'r', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                age = calculate_age(row['Дата народження'])
                category = categorize_age(age)
                employees_data.append({
                    'data': row,
                    'age': age,
                    'category': category
                })
        
        if not employees_data:
            print(f"Помилка: файл {csv_filename} порожній або має неправильний формат!")
            return
        
    except Exception as e:
        print(f"Помилка при відкритті файлу CSV: {e}")
        return
    
    # Створення XLSX файлу
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Створення аркушів
        sheets = {
            'all': wb.create_sheet('all'),
            'younger_18': wb.create_sheet('younger_18'),
            '18-45': wb.create_sheet('18-45'),
            '45-70': wb.create_sheet('45-70'),
            'older_70': wb.create_sheet('older_70')
        }
        
        # Заповнення аркуша "all"
        all_sheet = sheets['all']
        headers_all = list(employees_data[0]['data'].keys())
        all_sheet.append(headers_all)
        
        for emp in employees_data:
            all_sheet.append(list(emp['data'].values()))
        
        style_header(all_sheet)
        adjust_column_width(all_sheet)
        
        # Заповнення аркушів за віковими категоріями
        headers_age = ['№', 'Прізвище', 'Ім\'я', 'По батькові', 'Дата народження', 'Вік']
        
        for category_name, sheet in sheets.items():
            if category_name == 'all':
                continue
            
            sheet.append(headers_age)
            
            # Фільтрація даних за категорією
            filtered_data = [emp for emp in employees_data if emp['category'] == category_name]
            
            for idx, emp in enumerate(filtered_data, start=1):
                row_data = [
                    idx,
                    emp['data']['Прізвище'],
                    emp['data']['Ім\'я'],
                    emp['data']['По батькові'],
                    emp['data']['Дата народження'],
                    emp['age']
                ]
                sheet.append(row_data)
            
            style_header(sheet)
            adjust_column_width(sheet)
        
        # Збереження файлу
        wb.save(xlsx_filename)
        
        print("Ok")
        print(f"Файл {xlsx_filename} успішно створено!")
        print(f"Аркуш 'all': {len(employees_data)} записів")
        for category in ['younger_18', '18-45', '45-70', 'older_70']:
            count = len([e for e in employees_data if e['category'] == category])
            print(f"Аркуш '{category}': {count} записів")
        
    except Exception as e:
        print(f"Помилка: неможливість створення XLSX файлу - {e}")

if __name__ == "__main__":
    create_xlsx_from_csv('employees.csv', 'employees.xlsx')